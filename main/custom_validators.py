import csv
import openpyxl
from django.core.exceptions import ValidationError
from django.core.validators import EmailValidator
from django.utils.deconstruct import deconstructible
from django.utils.translation import gettext_lazy as _


@deconstructible
class EsiEmailValidator(EmailValidator):

    def validate_domain_part(self, domain_part):
        return False

    def __eq__(self, other):
        return isinstance(other, EsiEmailValidator) and super().__eq__(other)


def validate_emails_in_file(file):
    email_validator = EsiEmailValidator(allowlist=['esi.dz'],
                                        message='Enter a valid "@esi.dz" email address.')

    invalid_emails = []
    emails = []
    if str(file).endswith('.csv'):
        data = csv.reader(file)
        header = next(data)

        for row in data:
            emails.append(row[0])
    else:
        sheet = openpyxl.load_workbook(file)
        sheet = sheet.active

        # for row in range(1, sheet.max_row):
        #     for col in sheet.iter_cols(0):
        #         emails.append(col[row].value)

        for row in range(2, sheet.max_row + 1):
            emails.append(sheet.cell(row=row, column=1).value)

    for email in emails:
        if email:  # empty cases return None
            try:
                email_validator(email)
            except ValidationError:
                invalid_emails.append(email)

    file.seek(0)

    if invalid_emails:
        raise ValidationError([
            ValidationError(_('Invalid email: %(email)s'),
                            params={'email': email})
            for email in invalid_emails
        ])


def validate_file_size(file):
    filesize = file.size

    if filesize > 10485760*5:
        raise ValidationError("The maximum file size that can be uploaded is 10MB")
